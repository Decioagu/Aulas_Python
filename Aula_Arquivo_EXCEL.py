# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

'''
    "openpyxl" é uma biblioteca open-source escrita em Python utilizada para ler, 
    escrever e manipular arquivos do Microsoft Excel. Ela oferece uma forma fácil 
    e eficiente de automatizar tarefas relacionadas a planilhas.

    Principais funcionalidades do openpyxl:

    Leitura e escrita de arquivos Excel:
        # Suporta formatos .xlsx, .xlsm, .xltx e .xltm.
        # Carregamento e salvamento de planilhas do Excel.
    Manipulação de planilhas:
        # Criação, exclusão e renomeação de planilhas.
        # Acesso e iteração sobre células, linhas e colunas.
        # Adição, remoção e atualização de dados nas células.
    Formatação de planilhas:
        # Aplicação de estilos de fonte, cor, bordas e preenchimento.
        # Alinhamento de texto e números.
        # Formatação condicional de células.
    Fórmulas e funções:
        # Inserção de fórmulas do Excel nas células.
        # Acesso e modificação de fórmulas existentes.
        # Utilização de funções do Excel dentro das fórmulas. 
    Gráficos e tabelas dinâmicas:
        # Criação de diversos tipos de gráficos a partir dos dados da planilha.
        # Inserção de tabelas dinâmicas para análise e filtragem de dados.
'''

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet 

# caminho do arquivo
ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'estudantes.xlsx'

workbook = Workbook() # criar arquivo vazio do Excel

# ================ ADICIONAR NOME NA PLANINHA ================
# Nome para a planilha
sheet_name = 'Minha planilha'

# Criamos a planilha
workbook.create_sheet(sheet_name, 0) # alterar ordem (nome, posição)

# Selecionou a planilha (worksheet: Worksheet = tipagem)
worksheet: Worksheet = workbook[sheet_name] # planinha ativa ('Minha planilha')

# Remover uma planilha pelo nome
# workbook.remove(workbook['Sheet'])
# ============================ OU ==============================
# worksheet: Worksheet = workbook.active # planinha padrão ('Sheet')
# ==============================================================

print('===================================================================')
# ver nome das planinhas
for i in workbook:
    print(f'Nome planinha = {i}')

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')    # (linha, coluna, valor)
worksheet.cell(1, 2, 'Idade')   # (linha, coluna, valor)
worksheet.cell(1, 3, 'Nota')    # (linha, coluna, valor)

# lista de objetos
estudantes = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

print('===================================================================')
# ================================ TESTE DIDÁTICO ============================== 
for i, estudante_linha in enumerate(estudantes, start=2): # "start" manipula inicio da linha
    for j, estudante_coluna in enumerate(estudante_linha, start=1): # "start" manipula inicio da coluna
        worksheet.cell(i, j, estudante_coluna) # (linha, coluna, valor)
        print(f'linha = {i} |',f'coluna = {j} |', f'salvar = {estudante_coluna}')

# ==================================== OU ======================================= 

print('===================================================================')

# adicionar lista de objetos
for student in estudantes:
    worksheet.append(student) # adicionar como lista
    print(student)
# ===============================================================================

workbook.save(WORKBOOK_PATH) # salvar arquivo

#-----------------------------------------------------------------------------

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'estudantes.xlsx'

# ===================================================
# Caminho para leitura
workbook: Workbook = load_workbook(WORKBOOK_PATH) # Carregando um arquivo do excel
# Nome para a planilha
sheet_name = 'Minha planilha'
# Selecionou a planilha (worksheet: Worksheet = tipagem)
worksheet: Worksheet = workbook[sheet_name] # planinha ativa ('Minha planilha')
# ===================================================
print('===================================================================')
print(f"Ver valor da célula A9 no Excel = {worksheet['A1'].value}: {worksheet['A9'].value}")
print(f"Ver valor da célula B9 no Excel = {worksheet['B1'].value}: {worksheet['B9'].value}")
print(f"Ver valor da célula C9 no Excel = {worksheet['C1'].value}: {worksheet['C9'].value}")
print('===================================================================')
# ===================================================
row: tuple[Cell]
for linha in worksheet.iter_rows(min_row=6): # "min_row" = Ver linha inicial (não obrigatório)
    for coluna in linha:
        print(coluna.value, end='\t')

        # alterar valor da célula
        if coluna.value == 'Alberto':
                                               # "coluna.row" = valor numérico da linha
            worksheet.cell(coluna.row, 3, 9.9) # worksheet.cell(coluna.row, coluna, valor)
    print()
# ===================================================
print('===================================================================')
print(f"Ver valor da célula A9 no Excel = {worksheet['A1'].value}: {worksheet['A9'].value}")
print(f"Ver valor da célula B9 no Excel = {worksheet['B1'].value}: {worksheet['B9'].value}")
print(f"Ver valor da célula C9 no Excel = {worksheet['C1'].value}: {worksheet['C9'].value}")
print('===================================================================')

workbook.save(WORKBOOK_PATH)